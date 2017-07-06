from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.forms.formsets import formset_factory
from django.utils import timezone
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from .models import Element, Order, Product
from .forms import ElementForm, OrderForm, ProductForm


def init(request):

    return render(request, 'blog/init.html')


def element_list(request):

    elements = Element.objects.filter(created_date__lte=timezone.now()).order_by
    ('created_date')

    return render(request, 'blog/element_list.html', {'elements': elements})


def element_detail(request, pk):
    element = get_object_or_404(Element, pk=pk)

    return render(request, 'blog/element_detail.html', {'element': element})


@login_required
def element_new(request):
    if request.method == "POST":
        form = ElementForm(request.POST)
        if form.is_valid():
            element = form.save(commit=False)
            element.author = request.user
            # element.published_date = timezone.now()
            element.save()
            return redirect('blog:element_detail', pk=element.pk)
    else:
        form = ElementForm()
    return render(request, 'blog/element_edit.html', {'form': form})


@login_required
def element_edit(request, pk):
    element = get_object_or_404(Element, pk=pk)
    if request.method == "POST":
        form = ElementForm(data=request.POST, instance=element)
        if form.is_valid():
            element = form.save(commit=False)
            element.author = request.user
            element.save()
            return redirect('blog:element_detail', pk=element.pk)
    else:
        form = ElementForm(instance=element)
    return render(request, 'blog/element_edit.html', {'form': form})


@login_required
def element_remove(request, pk):
    element = get_object_or_404(Element, pk=pk)
    element.delete()
    return redirect('blog:element_list')


def order_list(request):

    orders = Order.objects.filter(created_date__lte=timezone.now()).order_by
    ('created_date')

    return render(request, 'blog/order_list.html', {'orders': orders})


def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk).order_by
    ('created_date')

    return render(request, 'blog/order_detail.html', {'order': order, 'products': products})


@login_required
def order_new(request):
    if request.method == "POST":
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.author = request.user
            order.save()
            return redirect('blog:order_new_next', pk=order.pk)
    else:
        form = OrderForm()
    return render(request, 'blog/order_new.html', {'form': form})


def setBordersCell(hoja):

    border_TopBottomThin = Border(top=Side(style='thin'),
                                  bottom=Side(style='thin'))

    border_RightTopBottomThin = Border(right=Side(style='thin'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='thin'))

    border_TopThin = Border(top=Side(style='thin'))

    border_BottomThin = Border(bottom=Side(style='thin'))

    border_TopThinBottomDouble = Border(top=Side(style='thin'),
                                        bottom=Side(style='double'))

    border_RightTopBottomMedium = Border(right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

    hoja.cell('D6').border = border_TopBottomThin
    hoja.cell('E6').border = border_RightTopBottomThin
    hoja.cell('D7').border = border_TopBottomThin
    hoja.cell('E7').border = border_RightTopBottomThin
    hoja.cell('D8').border = border_TopBottomThin
    hoja.cell('E8').border = border_RightTopBottomThin
    hoja.cell('D9').border = border_TopBottomThin
    hoja.cell('E9').border = border_RightTopBottomThin
    hoja.cell('D10').border = border_TopBottomThin
    hoja.cell('E10').border = border_RightTopBottomThin
    hoja.cell('D11').border = border_TopBottomThin
    hoja.cell('E11').border = border_RightTopBottomThin
    hoja.cell('D12').border = border_TopBottomThin
    hoja.cell('E12').border = border_RightTopBottomThin

    hoja.cell('G12').border = border_RightTopBottomMedium

    hoja.cell('D14').border = border_TopBottomThin
    hoja.cell('E14').border = border_TopBottomThin
    hoja.cell('F14').border = border_TopBottomThin
    hoja.cell('G14').border = border_TopBottomThin

    hoja.cell('D18').border = border_TopBottomThin
    hoja.cell('E18').border = border_TopBottomThin
    hoja.cell('F18').border = border_TopBottomThin
    hoja.cell('G18').border = border_TopBottomThin
    hoja.cell('H18').border = border_RightTopBottomThin

    hoja.cell('G30').border = border_RightTopBottomThin

    hoja.cell('B36').border = border_TopThin
    hoja.cell('C36').border = border_TopThin
    hoja.cell('D36').border = border_TopThin
    hoja.cell('E36').border = border_TopThin

    hoja.cell('B37').border = border_BottomThin
    hoja.cell('C37').border = border_BottomThin
    hoja.cell('D37').border = border_BottomThin
    hoja.cell('E37').border = border_BottomThin
    hoja.cell('F37').border = border_BottomThin
    hoja.cell('G37').border = border_BottomThin

    for num in range(38, 65):
        numStr = str(num)
        hoja.cell('B' + numStr).border = border_TopBottomThin
        hoja.cell('C' + numStr).border = border_TopBottomThin
        hoja.cell('D' + numStr).border = border_TopBottomThin
        hoja.cell('E' + numStr).border = border_TopBottomThin
        hoja.cell('F' + numStr).border = border_TopBottomThin
        hoja.cell('G' + numStr).border = border_RightTopBottomThin

    hoja.cell('I67').border = border_TopBottomThin
    hoja.cell('I68').border = border_TopBottomThin
    hoja.cell('I69').border = border_TopThinBottomDouble

    img = Image('blog/static/images/icn2.png')
    # img.anchor(hoja.cell('H2'))
    hoja.add_image(img, 'H2')
    return hoja


@login_required
def order_new_next(request, pk):
    order = get_object_or_404(Order, pk=pk)
    ProductFormSet = formset_factory(ProductForm, extra=order.number_product)
    if request.method == "POST":
        product = ProductForm()
        formset = ProductFormSet(request.POST)
        # doc = openpyxl.load_workbook('blog/formulariosPedidos/Formulario_Pedido_v2.xlsx',
        #                              formatting_info=True)
        doc = openpyxl.load_workbook('blog/formulariosPedidos/Formulario_Pedido_v2.xlsx')
        doc.get_sheet_names()
        hoja = doc.get_sheet_by_name('Order Form')
        hoja['C6'] = order.researcher
        hoja['C7'] = order.budget.name
        hoja['C10'] = order.buy_type.name
        hoja['C12'] = order.payment_requirements.name
        hoja['C18'] = order.provider.name
        num = 38
        nameFile = "FP_" + order.name
        if (formset.is_valid()):
            for form in formset:
                product = form.cleaned_data
                product = form.save(commit=False)
                numString = str(num)
                hoja['A' + numString] = product.description
                hoja['H' + numString] = product.quantity
                hoja['I' + numString] = product.unit_price
                num = num + 1
                product.order = order
                product.save()

            hoja = setBordersCell(hoja)
            doc.save('blog/formulariosPedidos/' + nameFile + '.xlsx')

            return redirect('blog:order_detail', pk=order.pk)
    else:
        formset = ProductFormSet()
    return render(request, 'blog/order_new_next.html', {'formset': formset})


@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk).order_by('created_date')
    ProductFormSet = formset_factory(ProductForm)
    if request.method == "POST":
        order_form = OrderForm(data=request.POST, instance=order, prefix="orderForm")
        formset = ProductFormSet(data=request.POST, prefix="productForm")
        print(len(products))
        if order_form.is_valid() and formset.is_valid():
            order = order_form.save(commit=False)
            order.author = request.user
            order.save()
            for num in range(0, len(products)):
                product = formset[num].save(commit=False)
                product.order = order
                if products[num].description != product.description:
                    products[num].description = product.description

                if products[num].quantity != product.quantity:
                    products[num].quantity = product.quantity

                if products[num].unit_price != product.unit_price:
                    products[num].unit_price = product.unit_price

                products[num].save()
            return redirect('blog:order_detail', pk=order.pk)
    else:
        order_form = OrderForm(instance=order, prefix="orderForm")
        products = Product.objects.filter(order=order.pk).order_by('created_date')
        products_formset = ProductFormSet(initial=[{'description': form.description,
                                                    'quantity': form.quantity,
                                                    'unit_price': form.unit_price}
                                                   for form in products], prefix="productForm")
    return render(request, 'blog/order_edit.html', {'order_form': order_form,
                                                    'products_formset': products_formset})


@login_required
def order_remove(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order.delete()
    return redirect('blog:order_list')
